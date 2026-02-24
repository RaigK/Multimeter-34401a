"""
HP/Agilent 34401A Multimeter - Grafische Benutzeroberfläche
============================================================
Funktionen:
  - Gerätekonfiguration (Messfunktion, Bereich, Auflösung, Triggermodus)
  - Kontinuierliche Messung mit einstellbarem Intervall
  - Grafische Darstellung mit optionaler Autoskalierung
  - Datenspeicherung im Excel-Format (.xlsx)
  - Simulation wenn kein Gerät angeschlossen ist

Voraussetzungen:
  pip install pyvisa matplotlib openpyxl numpy tkinter
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import time
import datetime
import queue
import random
import math

import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
import numpy as np

try:
    import pyvisa
    PYVISA_AVAILABLE = True
except ImportError:
    PYVISA_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ─────────────────────────────────────────────────────────────────────────────
# Geräteschnittstelle
# ─────────────────────────────────────────────────────────────────────────────

class Multimeter34401A:
    """Abstraktion für HP/Agilent 34401A (GPIB, RS-232, USB-GPIB)."""

    FUNCTIONS = {
        "DC Spannung":  "VOLT:DC",
        "AC Spannung":  "VOLT:AC",
        "DC Strom":     "CURR:DC",
        "AC Strom":     "CURR:AC",
        "2W Widerstand":"RES",
        "4W Widerstand":"FRES",
        "Frequenz":     "FREQ",
        "Periode":      "PER",
        "Durchgang":    "CONT",
        "Diode":        "DIOD",
    }

    RANGES = {
        "DC Spannung":  ["AUTO", "100 mV", "1 V", "10 V", "100 V", "1000 V"],
        "AC Spannung":  ["AUTO", "100 mV", "1 V", "10 V", "100 V", "750 V"],
        "DC Strom":     ["AUTO", "10 mA", "100 mA", "1 A", "3 A"],
        "AC Strom":     ["AUTO", "1 A", "3 A"],
        "2W Widerstand":["AUTO", "100 Ω", "1 kΩ", "10 kΩ", "100 kΩ", "1 MΩ", "10 MΩ", "100 MΩ"],
        "4W Widerstand":["AUTO", "100 Ω", "1 kΩ", "10 kΩ", "100 kΩ", "1 MΩ", "10 MΩ"],
        "Frequenz":     ["AUTO"],
        "Periode":      ["AUTO"],
        "Durchgang":    ["–"],
        "Diode":        ["–"],
    }

    RESOLUTIONS = ["3½ Digit", "4½ Digit", "5½ Digit", "6½ Digit"]
    RES_MAP = {"3½ Digit": 0.001, "4½ Digit": 0.0001,
               "5½ Digit": 0.00001, "6½ Digit": 0.000001}

    NPLC_MAP = {"3½ Digit": 0.02, "4½ Digit": 0.2,
                "5½ Digit": 1, "6½ Digit": 10}

    UNITS = {
        "DC Spannung": "V", "AC Spannung": "V",
        "DC Strom": "A",    "AC Strom": "A",
        "2W Widerstand": "Ω","4W Widerstand": "Ω",
        "Frequenz": "Hz",   "Periode": "s",
        "Durchgang": "Ω",   "Diode": "V",
    }

    def __init__(self):
        self.instrument = None
        self.simulation = True
        self._sim_phase = 0.0
        self._sim_func = "DC Spannung"

    # ── Verbindung ──────────────────────────────────────────────────────────

    def list_resources(self):
        if not PYVISA_AVAILABLE:
            return []
        try:
            rm = pyvisa.ResourceManager()
            return list(rm.list_resources())
        except Exception:
            return []

    def connect(self, resource_string: str) -> bool:
        if not PYVISA_AVAILABLE or not resource_string:
            self.simulation = True
            return False
        try:
            rm = pyvisa.ResourceManager()
            self.instrument = rm.open_resource(resource_string)
            self.instrument.timeout = 5000
            idn = self.instrument.query("*IDN?")
            if "34401" not in idn and "34410" not in idn:
                raise ValueError(f"Unbekanntes Gerät: {idn}")
            self.instrument.write("*RST")
            self.instrument.write("*CLS")
            self.simulation = False
            return True
        except Exception as e:
            self.instrument = None
            self.simulation = True
            raise e

    def disconnect(self):
        if self.instrument:
            try:
                self.instrument.write("SYST:LOC")
                self.instrument.close()
            except Exception:
                pass
            self.instrument = None
        self.simulation = True

    # ── Konfiguration ───────────────────────────────────────────────────────

    def configure(self, function: str, range_str: str, resolution: str):
        func_cmd = self.FUNCTIONS.get(function, "VOLT:DC")
        nplc = self.NPLC_MAP.get(resolution, 1)
        self._sim_func = function

        if self.simulation:
            return

        range_val = "DEF" if range_str in ("AUTO", "–") else self._parse_range(range_str)
        self.instrument.write(f"CONF:{func_cmd} {range_val}")
        if function not in ("Durchgang", "Diode", "Frequenz", "Periode"):
            self.instrument.write(f"SENS:{func_cmd}:NPLC {nplc}")
        self.instrument.write("TRIG:SOUR IMM")
        self.instrument.write("TRIG:DEL:AUTO ON")
        self.instrument.write("SAMP:COUN 1")

    @staticmethod
    def _parse_range(s: str) -> str:
        s = s.replace(" ", "").upper()
        multipliers = {"MV": 1e-3, "V": 1, "MA": 1e-3, "A": 1,
                       "Ω": 1, "KΩ": 1e3, "MΩ": 1e6, "HZ": 1}
        for suffix, mult in sorted(multipliers.items(), key=lambda x: -len(x[0])):
            if s.endswith(suffix):
                try:
                    return str(float(s[:-len(suffix)]) * mult)
                except ValueError:
                    pass
        return "DEF"

    # ── Messung ─────────────────────────────────────────────────────────────

    def measure(self) -> float:
        if self.simulation:
            return self._simulate()
        try:
            result = self.instrument.query("READ?")
            return float(result.strip())
        except Exception:
            return float("nan")

    def _simulate(self) -> float:
        self._sim_phase += 0.1
        f = self._sim_func
        if "Spannung" in f:
            base = 5.0 if "DC" in f else 230.0
            noise = random.gauss(0, base * 0.002)
            val = base + noise
            if "AC" in f:
                val = abs(val)
        elif "Strom" in f:
            base = 0.1 if "DC" in f else 0.5
            val = base + random.gauss(0, base * 0.005)
        elif "Widerstand" in f:
            val = 1000.0 + random.gauss(0, 0.5)
        elif "Frequenz" in f:
            val = 50.0 + random.gauss(0, 0.01)
        elif "Periode" in f:
            val = 0.02 + random.gauss(0, 1e-6)
        else:
            val = random.gauss(0, 0.001)
        return val


# ─────────────────────────────────────────────────────────────────────────────
# Haupt-GUI
# ─────────────────────────────────────────────────────────────────────────────

class MultimeterApp(tk.Tk):

    PLOT_COLORS = {
        "Hintergrund": "#1e1e2e", "Achsen": "#313244",
        "Linie": "#89b4fa", "Gitter": "#45475a",
        "Text": "#cdd6f4", "Akzent": "#a6e3a1",
    }

    def __init__(self):
        super().__init__()
        self.title("HP/Agilent 34401A Multimeter – Messsystem")
        self.geometry("1280x820")
        self.configure(bg="#1e1e2e")
        self.resizable(True, True)

        self.dmm = Multimeter34401A()
        self.data_timestamps: list[float] = []
        self.data_values: list[float] = []
        self.running = False
        self.measure_thread = None
        self.data_queue = queue.Queue()

        self._build_style()
        self._build_layout()
        self._update_plot_loop()

    # ── Stil ────────────────────────────────────────────────────────────────

    def _build_style(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        bg, fg, acc = "#1e1e2e", "#cdd6f4", "#89b4fa"
        style.configure("TFrame", background=bg)
        style.configure("TLabelframe", background=bg, foreground=fg,
                        bordercolor="#45475a", relief="groove")
        style.configure("TLabelframe.Label", background=bg, foreground=acc,
                        font=("Segoe UI", 10, "bold"))
        style.configure("TLabel", background=bg, foreground=fg,
                        font=("Segoe UI", 9))
        style.configure("TButton", background="#313244", foreground=fg,
                        font=("Segoe UI", 9, "bold"), padding=5)
        style.map("TButton",
                  background=[("active", "#45475a"), ("pressed", "#585b70")])
        style.configure("Accent.TButton", background="#89b4fa", foreground=bg,
                        font=("Segoe UI", 9, "bold"), padding=5)
        style.map("Accent.TButton",
                  background=[("active", "#74c7ec"), ("pressed", "#89dceb")])
        style.configure("Stop.TButton", background="#f38ba8", foreground=bg,
                        font=("Segoe UI", 9, "bold"), padding=5)
        style.configure("TCombobox", fieldbackground="#313244",
                        background="#313244", foreground=fg,
                        selectbackground="#45475a", selectforeground=fg)
        style.configure("TEntry", fieldbackground="#313244", foreground=fg,
                        insertcolor=fg)
        style.configure("TCheckbutton", background=bg, foreground=fg,
                        font=("Segoe UI", 9))

    # ── Layout ──────────────────────────────────────────────────────────────

    def _build_layout(self):
        # Linke Seitenleiste
        sidebar = ttk.Frame(self, width=310)
        sidebar.pack(side=tk.LEFT, fill=tk.Y, padx=8, pady=8)
        sidebar.pack_propagate(False)

        self._build_connection_frame(sidebar)
        self._build_function_frame(sidebar)
        self._build_acquisition_frame(sidebar)
        self._build_data_frame(sidebar)
        self._build_control_frame(sidebar)
        self._build_status_bar(sidebar)

        # Rechter Bereich: Display + Plot
        right = ttk.Frame(self)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=8, pady=8)

        self._build_display(right)
        self._build_plot(right)

    # ── Verbindungsrahmen ───────────────────────────────────────────────────

    def _build_connection_frame(self, parent):
        frm = ttk.LabelFrame(parent, text=" 🔌 Verbindung ", padding=8)
        frm.pack(fill=tk.X, pady=(0, 6))

        row1 = ttk.Frame(frm)
        row1.pack(fill=tk.X)
        ttk.Label(row1, text="Schnittstelle:").pack(side=tk.LEFT)

        self.resource_var = tk.StringVar(value="SIMULATION")
        self.resource_combo = ttk.Combobox(row1, textvariable=self.resource_var,
                                           width=22, state="normal")
        self.resource_combo.pack(side=tk.LEFT, padx=(6, 4))

        ttk.Button(row1, text="🔍", width=3,
                   command=self._scan_resources).pack(side=tk.LEFT)

        row2 = ttk.Frame(frm)
        row2.pack(fill=tk.X, pady=(6, 0))
        self.btn_connect = ttk.Button(row2, text="Verbinden",
                                      style="Accent.TButton",
                                      command=self._toggle_connect)
        self.btn_connect.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.conn_status = ttk.Label(row2, text="● Getrennt",
                                     foreground="#f38ba8", padding=(8, 0))
        self.conn_status.pack(side=tk.LEFT)

    # ── Funktionsrahmen ─────────────────────────────────────────────────────

    def _build_function_frame(self, parent):
        frm = ttk.LabelFrame(parent, text=" ⚙ Gerätefunktionen ", padding=8)
        frm.pack(fill=tk.X, pady=(0, 6))

        ttk.Label(frm, text="Messfunktion:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.func_var = tk.StringVar(value="DC Spannung")
        func_cb = ttk.Combobox(frm, textvariable=self.func_var, width=20,
                                values=list(Multimeter34401A.FUNCTIONS.keys()),
                                state="readonly")
        func_cb.grid(row=0, column=1, padx=(6, 0), pady=2, sticky=tk.EW)
        func_cb.bind("<<ComboboxSelected>>", self._on_function_change)

        ttk.Label(frm, text="Messbereich:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.range_var = tk.StringVar(value="AUTO")
        self.range_cb = ttk.Combobox(frm, textvariable=self.range_var, width=20,
                                     state="readonly")
        self.range_cb.grid(row=1, column=1, padx=(6, 0), pady=2, sticky=tk.EW)

        ttk.Label(frm, text="Auflösung:").grid(row=2, column=0, sticky=tk.W, pady=2)
        self.res_var = tk.StringVar(value="5½ Digit")
        res_cb = ttk.Combobox(frm, textvariable=self.res_var, width=20,
                               values=Multimeter34401A.RESOLUTIONS, state="readonly")
        res_cb.grid(row=2, column=1, padx=(6, 0), pady=2, sticky=tk.EW)

        frm.columnconfigure(1, weight=1)
        self._on_function_change()

    # ── Aufnahmerahmen ──────────────────────────────────────────────────────

    def _build_acquisition_frame(self, parent):
        frm = ttk.LabelFrame(parent, text=" ⏱ Aufnahme ", padding=8)
        frm.pack(fill=tk.X, pady=(0, 6))

        ttk.Label(frm, text="Intervall (ms):").grid(row=0, column=0,
                                                      sticky=tk.W, pady=2)
        self.interval_var = tk.StringVar(value="500")
        ttk.Entry(frm, textvariable=self.interval_var, width=10).grid(
            row=0, column=1, padx=(6, 0), pady=2, sticky=tk.EW)

        ttk.Label(frm, text="Max. Punkte:").grid(row=1, column=0,
                                                   sticky=tk.W, pady=2)
        self.maxpts_var = tk.StringVar(value="1000")
        ttk.Entry(frm, textvariable=self.maxpts_var, width=10).grid(
            row=1, column=1, padx=(6, 0), pady=2, sticky=tk.EW)

        self.autoscale_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(frm, text="Autoskalierung", variable=self.autoscale_var).grid(
            row=2, column=0, columnspan=2, sticky=tk.W, pady=2)

        self.statistics_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(frm, text="Statistik anzeigen",
                        variable=self.statistics_var).grid(
            row=3, column=0, columnspan=2, sticky=tk.W, pady=2)

        frm.columnconfigure(1, weight=1)

    # ── Datenrahmen ─────────────────────────────────────────────────────────

    def _build_data_frame(self, parent):
        frm = ttk.LabelFrame(parent, text=" 💾 Datei ", padding=8)
        frm.pack(fill=tk.X, pady=(0, 6))

        ttk.Label(frm, text="Dateiname (.xlsx):").pack(anchor=tk.W)

        row = ttk.Frame(frm)
        row.pack(fill=tk.X, pady=(4, 0))
        self.filename_var = tk.StringVar(
            value=f"Messung_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        ttk.Entry(row, textvariable=self.filename_var).pack(side=tk.LEFT,
                                                             fill=tk.X, expand=True)
        ttk.Button(row, text="📁", width=3,
                   command=self._browse_file).pack(side=tk.LEFT, padx=(4, 0))

        self.btn_save = ttk.Button(frm, text="💾 In Excel speichern",
                                   command=self._save_excel, state="disabled")
        self.btn_save.pack(fill=tk.X, pady=(6, 0))

        self.btn_clear = ttk.Button(frm, text="🗑 Daten löschen",
                                    command=self._clear_data)
        self.btn_clear.pack(fill=tk.X, pady=(4, 0))

    # ── Steuerrahmen ────────────────────────────────────────────────────────

    def _build_control_frame(self, parent):
        frm = ttk.Frame(parent)
        frm.pack(fill=tk.X, pady=(0, 6))

        self.btn_start = ttk.Button(frm, text="▶  Messung starten",
                                    style="Accent.TButton",
                                    command=self._start_measurement)
        self.btn_start.pack(fill=tk.X, pady=(0, 4))

        self.btn_stop = ttk.Button(frm, text="⏹  Messung stoppen",
                                   style="Stop.TButton",
                                   command=self._stop_measurement,
                                   state="disabled")
        self.btn_stop.pack(fill=tk.X)

    # ── Statusleiste ────────────────────────────────────────────────────────

    def _build_status_bar(self, parent):
        self.status_var = tk.StringVar(value="Bereit – Simulationsmodus")
        lbl = ttk.Label(parent, textvariable=self.status_var,
                        foreground="#a6e3a1", wraplength=290,
                        font=("Segoe UI", 8))
        lbl.pack(anchor=tk.W, pady=(4, 0))

    # ── Großes Display ──────────────────────────────────────────────────────

    def _build_display(self, parent):
        disp_frame = tk.Frame(parent, bg="#1e1e2e", height=100)
        disp_frame.pack(fill=tk.X, pady=(0, 6))
        disp_frame.pack_propagate(False)

        inner = tk.Frame(disp_frame, bg="#11111b", relief="sunken", bd=2)
        inner.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

        self.display_value = tk.StringVar(value="- - - - - -")
        self.display_unit = tk.StringVar(value="V")
        self.display_func = tk.StringVar(value="DC Spannung")

        tk.Label(inner, textvariable=self.display_func,
                 bg="#11111b", fg="#585b70",
                 font=("Courier New", 11)).pack(anchor=tk.W, padx=12, pady=(6, 0))

        val_row = tk.Frame(inner, bg="#11111b")
        val_row.pack(fill=tk.X, expand=True)

        tk.Label(val_row, textvariable=self.display_value,
                 bg="#11111b", fg="#a6e3a1",
                 font=("Courier New", 36, "bold")).pack(side=tk.LEFT, padx=12)
        tk.Label(val_row, textvariable=self.display_unit,
                 bg="#11111b", fg="#89b4fa",
                 font=("Courier New", 22)).pack(side=tk.LEFT, pady=(8, 0))

        self.sim_indicator = tk.Label(inner, text="SIM", bg="#11111b",
                                      fg="#fab387", font=("Courier New", 9))
        self.sim_indicator.pack(anchor=tk.NE, padx=12)

    # ── Diagramm ────────────────────────────────────────────────────────────

    def _build_plot(self, parent):
        plot_frame = ttk.Frame(parent)
        plot_frame.pack(fill=tk.BOTH, expand=True)

        c = self.PLOT_COLORS
        self.fig = Figure(figsize=(8, 4.5), facecolor=c["Hintergrund"])
        self.ax = self.fig.add_subplot(111)
        self._style_axes()
        self.line, = self.ax.plot([], [], color=c["Linie"], linewidth=1.5,
                                  antialiased=True)
        self.fig.tight_layout(pad=1.2)

        self.canvas = FigureCanvasTkAgg(self.fig, master=plot_frame)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        toolbar_frame = ttk.Frame(plot_frame)
        toolbar_frame.pack(fill=tk.X)
        self.toolbar = NavigationToolbar2Tk(self.canvas, toolbar_frame)
        self.toolbar.update()

        # Statistik-Textfeld
        self.stat_text = self.ax.text(
            0.01, 0.98, "", transform=self.ax.transAxes,
            verticalalignment="top", fontfamily="monospace", fontsize=8,
            color=self.PLOT_COLORS["Akzent"],
            bbox=dict(boxstyle="round,pad=0.3", facecolor="#181825",
                      edgecolor="#45475a", alpha=0.8))

    def _style_axes(self):
        c = self.PLOT_COLORS
        self.ax.set_facecolor(c["Achsen"])
        self.ax.tick_params(colors=c["Text"], labelsize=8)
        self.ax.xaxis.label.set_color(c["Text"])
        self.ax.yaxis.label.set_color(c["Text"])
        for spine in self.ax.spines.values():
            spine.set_edgecolor(c["Gitter"])
        self.ax.grid(True, color=c["Gitter"], linewidth=0.5, alpha=0.6)
        self.ax.set_xlabel("Zeit (s)", color=c["Text"], fontsize=9)
        self.ax.set_ylabel("Messwert", color=c["Text"], fontsize=9)
        self.ax.set_title("Messverlauf", color=c["Text"], fontsize=10)

    # ── Callbacks ───────────────────────────────────────────────────────────

    def _on_function_change(self, event=None):
        func = self.func_var.get()
        ranges = Multimeter34401A.RANGES.get(func, ["AUTO"])
        self.range_cb["values"] = ranges
        self.range_var.set(ranges[0])
        unit = Multimeter34401A.UNITS.get(func, "")
        self.display_unit.set(unit)
        self.display_func.set(func)
        self.ax.set_ylabel(f"Messwert ({unit})", color=self.PLOT_COLORS["Text"],
                           fontsize=9)
        self.canvas.draw_idle()

    def _scan_resources(self):
        resources = self.dmm.list_resources()
        vals = ["SIMULATION"] + resources
        self.resource_combo["values"] = vals
        self.status_var.set(
            f"{len(resources)} Gerät(e) gefunden" if resources
            else "Keine Geräte gefunden – Simulationsmodus verfügbar")

    def _toggle_connect(self):
        if not self.dmm.simulation:
            self.dmm.disconnect()
            self.conn_status.configure(text="● Getrennt", foreground="#f38ba8")
            self.btn_connect.configure(text="Verbinden", style="Accent.TButton")
            self.sim_indicator.configure(text="SIM")
            self.status_var.set("Getrennt")
        else:
            res = self.resource_var.get()
            if res == "SIMULATION":
                self.conn_status.configure(text="● Simulation", foreground="#fab387")
                self.sim_indicator.configure(text="SIM")
                self.status_var.set("Simulationsmodus aktiv")
            else:
                try:
                    self.dmm.connect(res)
                    self.conn_status.configure(text="● Verbunden",
                                               foreground="#a6e3a1")
                    self.btn_connect.configure(text="Trennen", style="Stop.TButton")
                    self.sim_indicator.configure(text="")
                    self.status_var.set(f"Verbunden: {res}")
                except Exception as e:
                    messagebox.showerror("Verbindungsfehler", str(e))
                    self.status_var.set(f"Fehler: {e}")

    def _start_measurement(self):
        try:
            interval_ms = int(self.interval_var.get())
            if interval_ms < 50:
                raise ValueError("Mindestintervall 50 ms")
        except ValueError as e:
            messagebox.showerror("Eingabefehler", str(e))
            return

        # Gerät konfigurieren
        self.dmm.configure(self.func_var.get(),
                           self.range_var.get(),
                           self.res_var.get())

        self.running = True
        self.btn_start.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self.btn_save.configure(state="disabled")
        self.status_var.set("Messung läuft …")

        self.measure_thread = threading.Thread(
            target=self._measure_loop, args=(interval_ms,), daemon=True)
        self.measure_thread.start()

    def _stop_measurement(self):
        self.running = False
        self.btn_start.configure(state="normal")
        self.btn_stop.configure(state="disabled")
        if self.data_values:
            self.btn_save.configure(state="normal")
        self.status_var.set(
            f"Messung gestoppt – {len(self.data_values)} Punkte aufgenommen")

    def _measure_loop(self, interval_ms: int):
        t0 = time.time()
        while self.running:
            t_start = time.time()
            val = self.dmm.measure()
            elapsed = time.time() - t0
            self.data_queue.put((elapsed, val))
            sleep_time = (interval_ms / 1000.0) - (time.time() - t_start)
            if sleep_time > 0:
                time.sleep(sleep_time)

    def _update_plot_loop(self):
        """Wird im Haupt-Thread aufgerufen; liest aus Queue und aktualisiert Plot."""
        changed = False
        try:
            maxpts = int(self.maxpts_var.get())
        except ValueError:
            maxpts = 1000

        while not self.data_queue.empty():
            t, v = self.data_queue.get_nowait()
            self.data_timestamps.append(t)
            self.data_values.append(v)
            changed = True

        if changed and self.data_values:
            # Begrenzen
            if len(self.data_values) > maxpts:
                self.data_timestamps = self.data_timestamps[-maxpts:]
                self.data_values = self.data_values[-maxpts:]

            v = self.data_values[-1]
            unit = Multimeter34401A.UNITS.get(self.func_var.get(), "")
            self.display_value.set(f"{v:>+14.7g}")

            # Plot
            self.line.set_data(self.data_timestamps, self.data_values)

            if self.autoscale_var.get():
                self.ax.relim()
                self.ax.autoscale_view()
            else:
                tmin, tmax = min(self.data_timestamps), max(self.data_timestamps)
                vmin, vmax = min(self.data_values), max(self.data_values)
                margin = (vmax - vmin) * 0.05 if vmax != vmin else abs(vmax) * 0.05
                self.ax.set_xlim(tmin, tmax if tmax > tmin else tmin + 1)
                self.ax.set_ylim(vmin - margin, vmax + margin)

            # Statistik
            if self.statistics_var.get() and len(self.data_values) >= 2:
                arr = np.array(self.data_values)
                txt = (f"n = {len(arr)}\n"
                       f"μ = {arr.mean():.6g} {unit}\n"
                       f"σ = {arr.std():.4g} {unit}\n"
                       f"min = {arr.min():.6g} {unit}\n"
                       f"max = {arr.max():.6g} {unit}")
                self.stat_text.set_text(txt)
            else:
                self.stat_text.set_text("")

            self.canvas.draw_idle()

        self.after(80, self._update_plot_loop)

    def _browse_file(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel-Dateien", "*.xlsx"), ("Alle Dateien", "*.*")],
            initialfile=self.filename_var.get())
        if path:
            self.filename_var.set(path)

    def _clear_data(self):
        if self.running:
            messagebox.showwarning("Warnung",
                                   "Messung zuerst stoppen, dann Daten löschen.")
            return
        self.data_timestamps.clear()
        self.data_values.clear()
        self.line.set_data([], [])
        self.ax.relim()
        self.ax.autoscale_view()
        self.stat_text.set_text("")
        self.canvas.draw_idle()
        self.display_value.set("- - - - - -")
        self.btn_save.configure(state="disabled")
        self.status_var.set("Daten gelöscht")

    # ── Excel-Export ────────────────────────────────────────────────────────

    def _save_excel(self):
        if not self.data_values:
            messagebox.showinfo("Keine Daten", "Es wurden keine Messdaten aufgenommen.")
            return
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Fehler", "Modul 'openpyxl' nicht installiert.\n"
                                           "Bitte: pip install openpyxl")
            return

        path = self.filename_var.get()
        if not path:
            messagebox.showwarning("Dateiname fehlt", "Bitte einen Dateinamen eingeben.")
            return

        try:
            self._write_excel(path)
            messagebox.showinfo("Gespeichert",
                                f"Datei erfolgreich gespeichert:\n{path}")
            self.status_var.set(f"Gespeichert: {path}")
        except Exception as e:
            messagebox.showerror("Speicherfehler", str(e))

    def _write_excel(self, path: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Messdaten"

        func = self.func_var.get()
        unit = Multimeter34401A.UNITS.get(func, "")
        now = datetime.datetime.now()

        # ── Kopf ────────────────────────────────────────────────────────────
        header_fill = PatternFill("solid", fgColor="1e3a5f")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        title_font  = Font(name="Calibri", bold=True, color="89B4FA", size=14)

        ws["A1"] = "HP/Agilent 34401A – Messung"
        ws["A1"].font = title_font
        ws.merge_cells("A1:E1")

        meta = [
            ("Datum / Zeit:", now.strftime("%Y-%m-%d %H:%M:%S")),
            ("Messfunktion:", func),
            ("Messbereich:", self.range_var.get()),
            ("Auflösung:", self.res_var.get()),
            ("Intervall (ms):", self.interval_var.get()),
            ("Anzahl Punkte:", len(self.data_values)),
        ]
        for i, (k, v) in enumerate(meta, start=2):
            ws.cell(row=i, column=1, value=k).font = Font(bold=True, color="74C7EC")
            ws.cell(row=i, column=2, value=str(v))

        # ── Spaltentitel ────────────────────────────────────────────────────
        header_row = len(meta) + 3
        cols = ["#", "Zeit (s)", f"Messwert ({unit})", "Datum / Zeit"]
        for col, txt in enumerate(cols, start=1):
            cell = ws.cell(row=header_row, column=col, value=txt)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # ── Daten ────────────────────────────────────────────────────────────
        t0 = self.data_timestamps[0]
        abs_t0 = now - datetime.timedelta(seconds=self.data_timestamps[-1])
        thin = Side(style="thin", color="45475a")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for i, (ts, val) in enumerate(
                zip(self.data_timestamps, self.data_values), start=1):
            row = header_row + i
            abs_time = abs_t0 + datetime.timedelta(seconds=ts)
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=round(ts - t0, 4))
            ws.cell(row=row, column=3, value=round(val, 9))
            ws.cell(row=row, column=4, value=abs_time.strftime("%H:%M:%S.%f")[:-3])
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = border
                if i % 2 == 0:
                    ws.cell(row=row, column=col).fill = PatternFill(
                        "solid", fgColor="1e1e2e")

        # ── Statistik ────────────────────────────────────────────────────────
        arr = np.array(self.data_values)
        stat_row = header_row + len(self.data_values) + 2
        stats = [
            ("Mittelwert (μ):", f"{arr.mean():.9g} {unit}"),
            ("Std.-Abw. (σ):", f"{arr.std():.6g} {unit}"),
            ("Minimum:",       f"{arr.min():.9g} {unit}"),
            ("Maximum:",       f"{arr.max():.9g} {unit}"),
            ("Peak-Peak:",     f"{arr.max()-arr.min():.6g} {unit}"),
        ]
        ws.cell(row=stat_row, column=1, value="Statistik").font = Font(
            bold=True, color="A6E3A1", size=11)
        for j, (k, v) in enumerate(stats, start=1):
            ws.cell(row=stat_row + j, column=1, value=k).font = Font(bold=True)
            ws.cell(row=stat_row + j, column=2, value=v)

        # ── Liniendiagramm ───────────────────────────────────────────────────
        ws2 = wb.create_sheet("Diagramm")
        # Daten für Chart in ws2 kopieren
        ws2["A1"] = "Zeit (s)"
        ws2["B1"] = f"Messwert ({unit})"
        for i, (ts, val) in enumerate(
                zip(self.data_timestamps, self.data_values), start=2):
            ws2.cell(row=i, column=1, value=round(ts - t0, 4))
            ws2.cell(row=i, column=2, value=round(val, 9))

        chart = LineChart()
        chart.title = f"{func} – Messverlauf"
        chart.style = 10
        chart.y_axis.title = f"Messwert ({unit})"
        chart.x_axis.title = "Zeit (s)"
        data_ref = Reference(ws2, min_col=2, min_row=1,
                             max_row=len(self.data_values) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.width = 20
        chart.height = 12
        ws2.add_chart(chart, "D2")

        # ── Spaltenbreiten anpassen ──────────────────────────────────────────
        for ws_obj in (ws, ws2):
            for col in ws_obj.columns:
                max_len = max(
                    (len(str(cell.value)) for cell in col if cell.value), default=0)
                ws_obj.column_dimensions[col[0].column_letter].width = min(
                    max_len + 4, 40)

        wb.save(path)


# ─────────────────────────────────────────────────────────────────────────────
# Einstiegspunkt
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    # Abhängigkeiten prüfen
    missing = []
    if not PYVISA_AVAILABLE:
        missing.append("pyvisa")
    if not OPENPYXL_AVAILABLE:
        missing.append("openpyxl")
    if missing:
        import sys
        print(f"[WARNUNG] Fehlende Module: {', '.join(missing)}")
        print(f"Bitte installieren: pip install {' '.join(missing)}")
        print("Das Programm läuft im eingeschränkten Modus weiter.\n")

    app = MultimeterApp()
    app.mainloop()
